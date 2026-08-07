#!/usr/bin/env python3
"""Valida, audita e atualiza somente o campo ``tom`` dos hinos no Firestore."""

from __future__ import annotations

import argparse
import csv
import json
import os
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path


LIVROS = {"CC": 581, "VM": 435}
TONS_VALIDOS = {
    "C", "Cm", "Db", "D", "Dm", "Eb", "E", "Em", "F", "Fm",
    "Gb", "G", "Gm", "Ab", "A", "Am", "Bb", "B", "Bm",
}


def normalizar_tom(valor: object) -> str:
    tom = str(valor or "").strip()
    if tom.upper() == "EB":
        return "Eb"
    if not tom:
        return ""
    return tom[0].upper() + tom[1:].lower()


def carregar_planilha(caminho: Path) -> dict[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as erro:
        raise SystemExit(
            "Pacote ausente. Instale com: python -m pip install openpyxl"
        ) from erro

    if not caminho.is_file():
        raise SystemExit(f"Planilha não encontrada: {caminho}")

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    if "Tonalidades" not in workbook.sheetnames:
        raise SystemExit("A planilha precisa conter a aba 'Tonalidades'.")
    sheet = workbook["Tonalidades"]
    cabecalho = [str(c.value or "").strip() for c in sheet[1][:3]]
    if cabecalho != ["LIVRO", "Num", "Tom"]:
        raise SystemExit(
            "Cabeçalho inesperado. Esperado: LIVRO | Num | Tom; "
            f"encontrado: {cabecalho}"
        )

    tons: dict[str, str] = {}
    numeros: dict[str, set[int]] = {livro: set() for livro in LIVROS}
    for linha, (livro_bruto, numero_bruto, tom_bruto) in enumerate(
        sheet.iter_rows(min_row=2, max_col=3, values_only=True), start=2
    ):
        livro = str(livro_bruto or "").strip().upper()
        if livro not in LIVROS:
            raise SystemExit(f"Linha {linha}: livro inválido: {livro_bruto!r}")
        if not isinstance(numero_bruto, (int, float)) or int(numero_bruto) != numero_bruto:
            raise SystemExit(f"Linha {linha}: número inválido: {numero_bruto!r}")
        numero = int(numero_bruto)
        if not 1 <= numero <= LIVROS[livro]:
            raise SystemExit(f"Linha {linha}: número fora da faixa de {livro}: {numero}")
        tom = normalizar_tom(tom_bruto)
        if tom not in TONS_VALIDOS:
            raise SystemExit(f"Linha {linha}: tonalidade inválida: {tom_bruto!r}")

        documento_id = f"{livro}_{numero:03d}"
        if documento_id in tons:
            raise SystemExit(f"Linha {linha}: hino duplicado: {documento_id}")
        tons[documento_id] = tom
        numeros[livro].add(numero)

    for livro, total in LIVROS.items():
        esperados = set(range(1, total + 1))
        if numeros[livro] != esperados:
            ausentes = sorted(esperados - numeros[livro])
            extras = sorted(numeros[livro] - esperados)
            raise SystemExit(f"Sequência inválida em {livro}; ausentes={ausentes}, extras={extras}")

    if len(tons) != sum(LIVROS.values()):
        raise SystemExit(f"Esperados 1016 hinos; encontrados {len(tons)}.")
    return tons


def conectar_firestore():
    if not os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
        raise SystemExit(
            "Defina GOOGLE_APPLICATION_CREDENTIALS com o caminho da conta de serviço."
        )
    try:
        import firebase_admin
        from firebase_admin import firestore
    except ImportError as erro:
        raise SystemExit(
            "Pacote ausente. Instale com: python -m pip install firebase-admin"
        ) from erro

    if not firebase_admin._apps:
        firebase_admin.initialize_app()
    return firestore.client()


def auditar(banco, colecao: str, tons: dict[str, str]) -> list[dict]:
    referencias = [banco.collection(colecao).document(doc_id) for doc_id in tons]
    documentos = list(banco.get_all(referencias))
    encontrados = {doc.id: doc for doc in documentos if doc.exists}
    relatorio: list[dict] = []

    for doc_id, novo_tom in tons.items():
        documento = encontrados.get(doc_id)
        if documento is None:
            relatorio.append({
                "id": doc_id, "titulo": "", "tom_atual": "",
                "tom_novo": novo_tom, "status": "AUSENTE",
            })
            continue
        dados = documento.to_dict() or {}
        tom_atual = str(dados.get("tom", "")).strip()
        status = "IGUAL" if tom_atual == novo_tom else "ALTERAR"
        if dados.get("id") != doc_id:
            status = "ID_INTERNO_DIVERGENTE"
        relatorio.append({
            "id": doc_id,
            "titulo": str(dados.get("titulo", "")),
            "tom_atual": tom_atual,
            "tom_novo": novo_tom,
            "status": status,
        })
    return relatorio


def salvar_csv(relatorio: list[dict], caminho: Path) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with caminho.open("w", encoding="utf-8-sig", newline="") as arquivo:
        escritor = csv.DictWriter(
            arquivo, fieldnames=["id", "titulo", "tom_atual", "tom_novo", "status"]
        )
        escritor.writeheader()
        escritor.writerows(relatorio)


def salvar_backup(relatorio: list[dict], caminho: Path, colecao: str) -> None:
    conteudo = {
        "colecao": colecao,
        "criadoEmUtc": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "tonsAnteriores": {item["id"]: item["tom_atual"] for item in relatorio},
    }
    caminho.write_text(json.dumps(conteudo, ensure_ascii=False, indent=2), encoding="utf-8")


def atualizar(banco, colecao: str, relatorio: list[dict]) -> int:
    alteracoes = [item for item in relatorio if item["status"] == "ALTERAR"]
    gravados = 0
    for inicio in range(0, len(alteracoes), 400):
        lote = banco.batch()
        parte = alteracoes[inicio: inicio + 400]
        for item in parte:
            referencia = banco.collection(colecao).document(item["id"])
            lote.update(referencia, {"tom": item["tom_novo"]})
        lote.commit()
        gravados += len(parte)
        print(f"Atualizados: {gravados}/{len(alteracoes)}")
    return gravados


def sincronizar_json(caminho: Path, tons: dict[str, str], compacto: bool) -> int:
    if not caminho.is_file():
        raise SystemExit(f"Arquivo local não encontrado: {caminho}")
    with caminho.open("r", encoding="utf-8") as arquivo:
        registros = json.load(arquivo)
    if not isinstance(registros, list):
        raise SystemExit(f"O arquivo não contém uma lista JSON: {caminho}")

    ids_encontrados: set[str] = set()
    alterados = 0
    for posicao, registro in enumerate(registros, start=1):
        if not isinstance(registro, dict):
            raise SystemExit(f"Item {posicao} inválido em {caminho}")
        doc_id = str(registro.get("id", ""))
        if doc_id in ids_encontrados:
            raise SystemExit(f"ID duplicado em {caminho}: {doc_id}")
        ids_encontrados.add(doc_id)
        if doc_id not in tons:
            raise SystemExit(f"ID inesperado em {caminho}: {doc_id}")
        if "tom" not in registro:
            raise SystemExit(f"Campo tom ausente em {caminho}: {doc_id}")
        if registro["tom"] != tons[doc_id]:
            registro["tom"] = tons[doc_id]
            alterados += 1

    ausentes = sorted(set(tons) - ids_encontrados)
    if ausentes:
        raise SystemExit(f"IDs ausentes em {caminho}: {ausentes[:10]}")

    temporario = caminho.with_suffix(caminho.suffix + ".tmp")
    with temporario.open("w", encoding="utf-8", newline="\n") as arquivo:
        if compacto:
            json.dump(registros, arquivo, ensure_ascii=False, separators=(",", ":"))
        else:
            json.dump(registros, arquivo, ensure_ascii=False, indent=2)
            arquivo.write("\n")
    temporario.replace(caminho)
    return alterados


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("planilha", type=Path, help="Arquivo TONS_CHECKED.xlsx")
    parser.add_argument("--colecao", default="hinos_v2", help="Padrão: hinos_v2")
    grupo = parser.add_mutually_exclusive_group()
    grupo.add_argument("--auditar", action="store_true", help="Compara com o Firestore sem gravar")
    grupo.add_argument("--executar", action="store_true", help="Atualiza somente o campo tom")
    grupo.add_argument(
        "--sincronizar-locais",
        action="store_true",
        help="Atualiza hinos_firestore.json e o índice local, sem acessar o Firebase",
    )
    parser.add_argument(
        "--confirmar", default="", help="Para executar, informe exatamente ATUALIZAR-TONS"
    )
    parser.add_argument("--relatorio", type=Path, default=Path("relatorio_tonalidades.csv"))
    parser.add_argument("--json-hinos", type=Path, default=Path("hinos_firestore.json"))
    parser.add_argument(
        "--indice", type=Path, default=Path("assets/indice_busca_hinos.json")
    )
    args = parser.parse_args()

    if "/" in args.colecao or "\\" in args.colecao or not args.colecao.strip():
        raise SystemExit("Informe somente o nome de uma coleção válida.")

    tons = carregar_planilha(args.planilha)
    contagem = Counter(tons.keys())
    assert all(valor == 1 for valor in contagem.values())
    print("PLANILHA VÁLIDA: 1016 hinos (CC 1–581 e VM 1–435).")
    print("Padronização aplicada: CC_009, EB -> Eb.")

    if args.sincronizar_locais:
        alterados_hinos = sincronizar_json(args.json_hinos, tons, compacto=False)
        alterados_indice = sincronizar_json(args.indice, tons, compacto=True)
        print(f"BASE LOCAL SINCRONIZADA: {alterados_hinos} tonalidades alteradas.")
        print(f"ÍNDICE LOCAL SINCRONIZADO: {alterados_indice} tonalidades alteradas.")
        print("Nenhuma conexão com o Firebase foi aberta.")
        return

    if not args.auditar and not args.executar:
        print("VALIDAÇÃO LOCAL: nenhuma conexão foi aberta e nada foi gravado.")
        print("Próximo passo seguro: execute novamente com --auditar.")
        return

    banco = conectar_firestore()
    relatorio = auditar(banco, args.colecao, tons)
    salvar_csv(relatorio, args.relatorio)
    totais = Counter(item["status"] for item in relatorio)
    print(f"Coleção auditada: {args.colecao}")
    print("Resultado: " + ", ".join(f"{chave}={valor}" for chave, valor in sorted(totais.items())))
    print(f"Relatório: {args.relatorio.resolve()}")

    bloqueios = [item for item in relatorio if item["status"] not in {"IGUAL", "ALTERAR"}]
    if bloqueios:
        raise SystemExit("ATUALIZAÇÃO BLOQUEADA: corrija os itens inconsistentes do relatório.")

    if not args.executar:
        print("AUDITORIA: nenhum documento foi modificado.")
        return
    if args.confirmar != "ATUALIZAR-TONS":
        raise SystemExit("Para gravar, acrescente: --confirmar ATUALIZAR-TONS")

    instante = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    backup = args.relatorio.with_name(f"backup_tons_{args.colecao}_{instante}.json")
    salvar_backup(relatorio, backup, args.colecao)
    print(f"Backup dos tons anteriores: {backup.resolve()}")
    total = atualizar(banco, args.colecao, relatorio)
    print(f"ATUALIZAÇÃO CONCLUÍDA: {total} tonalidades alteradas.")


if __name__ == "__main__":
    main()
